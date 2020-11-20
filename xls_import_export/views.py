import os
from Stock_Manager import app, db
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from datetime import datetime
from part_mng.models import Part
from part_mng.helpers import find_part_in_db
from package_mng.models import Package, AltPackage
from package_mng.helpers import package_in_db_get_id_or_none
from xls_import_export.form import FileUploadForm, MapTableColumnsForPartForm, MapTableColumnsForPackageForm
from xls_import_export.helpers import already_in_list
from flask import render_template, redirect, request, url_for, flash, send_from_directory

SELECTFIELD_NONE_SELECTED = 100

# #############################################################################
#
#                       D O W N L O A D    D A T A 
#
# #############################################################################


# creates a workbook in memory. Writes a header line based on the column
# names of the given database_table. Then querys all items in the database
# and stores their values in the workbook. Saves file when done.
# unfortunately this does not work if data based on table relationships 
# is needed (as only the id of the foreignKey will be stored)
def create_download_file(database_orm_obj, prefix=""):
    book = Workbook()
    sheet = book.active # currently active worksheet (default 0)

    col_keys = database_orm_obj.__table__.columns.keys()
    sheet.append(col_keys)

    for item in database_orm_obj.query.all():
        row = []
        item=item.__dict__
        for col_key in col_keys:    # ??? iterate over columns
            row.append(item[col_key])   # ??? add table row's value for this column
        sheet.append(row)             # append all the tables row values to excel file
    datetimestr = datetime.utcnow().strftime("%Y-%m-%d_%H-%M")
    sheet.title = prefix + datetimestr
    book.save("generated/%s.xlsx" % sheet.title)
    return sheet.title+'.xlsx'


def create_download_file_finish(book, prefix):
    datetimestr = datetime.utcnow().strftime("%Y-%m-%d_%H-%M")
    sheet = book.active
    sheet.title = prefix + datetimestr
    book.save("generated/%s.xlsx" % sheet.title)
    return sheet.title+'.xlsx'

@app.route('/download/part')
def part_downloadXls():
    book = Workbook()
    sheet = book.active # currently active worksheet (default 0)
    row = ('Name','Manufacturer','Ordering Code','Package')
    sheet.append(row)
    for part in Part.query.all():
        row = ( part.name,
                part.manufacturer,
                part.ordering_code,
                part.package.name
              )
        sheet.append(row)
    filename = create_download_file_finish(book = book, prefix='part_')
    print app.config['GENERATED_DOWNLOADS_FOLDER']+filename
    return send_from_directory(directory=app.config['GENERATED_DOWNLOADS_FOLDER'],
                               filename=filename, as_attachment=True)

@app.route('/download/package')
def package_downloadXls():
    book = Workbook()
    sheet = book.active # currently active worksheet (default 0)
    row = ('Name','Pin Count','Pitch','Length', 'Width','Height', 'Alternative Names')
    sheet.append(row)
    for package in Package.query.all():
        alt_name_list = []
        for alt in package.alt_names:
            alt_name_list.append(alt.name)
        alt_names_string = ", ".join(alt_name_list)
        row = ( package.name,
                package.pin_count,
                package.pitch,
                package.length,
                package.width,
                package.height,
                alt_names_string
              )
        sheet.append(row)
    filename = create_download_file_finish(book=book, prefix='package_')
    return send_from_directory(directory=app.config['GENERATED_DOWNLOADS_FOLDER'],
                               filename=filename, as_attachment=True)


# #############################################################################
#
#                         U P L O A D    D A T A 
#
# #############################################################################

def getTableHeaders(filepath):
    wb = load_workbook(filename=filepath, read_only=True)
    sheet = wb.active
    col_headings = []
    index = 0
    dim = sheet.calculate_dimension()
    dim = dim.split(':')
    xy = coordinate_from_string(dim[1])
    max_col = column_index_from_string(xy[0])
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=max_col):
        for cell in row:
            index += 1
            if cell.value != None:
                col_headings.append((index, cell.value))
    return col_headings

@app.route('/file/upload', methods=('GET', 'POST'))
def file_upload():
    form = FileUploadForm()
    if form.validate_on_submit():
        table = request.files['import_file']
        content_type = form.import_type.data
        filename = "uploads/"+content_type+"/"+secure_filename(table.filename)
        table.save(filename)
        flash('File uploaded', 'success')
        return redirect(url_for('file_process', content_type=content_type, filename=secure_filename(table.filename)))
    return render_template('xls_import_export/upload.html', form=form)

@app.route('/file/prepare/<content_type>/<filename>', methods=['GET','POST'])
def file_process(content_type, filename):
    filepath = 'uploads/'+content_type+'/'+filename
    choices = getTableHeaders(filepath)
    choices_allow_None = [(SELECTFIELD_NONE_SELECTED,"---")]+choices

    if content_type == 'part':
        form = MapTableColumnsForPartForm()
        form.manufacturer.choices = choices_allow_None
        form.order_code.choices = choices
        form.part_name.choices = choices_allow_None
        form.package.choices = choices
        form.description.choices = choices_allow_None
        if form.validate_on_submit():
            import_parts(filepath=filepath, form=form)
    elif content_type == 'package':
        form = MapTableColumnsForPackageForm()
        form.name.choices = choices
        form.pin_count.choices = choices_allow_None
        form.pitch.choices = choices_allow_None
        form.width.choices = choices_allow_None
        form.length.choices = choices_allow_None
        form.height.choices = choices_allow_None
        form.alt_package.choices = choices_allow_None
        if form.validate_on_submit():
            import_packages(filepath=filepath, form=form)
    elif content_type == 'assembly':
        form = None
    else:
        form = None
    return render_template('xls_import_export/map_columns.html', form=form, content_type=content_type, filename = filename)

def import_parts(filepath, form):
    book = load_workbook(filename=filepath, read_only=True)
    sheet = book.active
    obj_to_add_to_db = create_partlist_from_part_sheet(form=form, sheet=sheet, max_row=1000)
    addToDatabase(obj_to_add_to_db)

def create_partlist_from_part_sheet(form, sheet, max_row):
    obj_to_add_to_db = []
    identifiers_to_add = []
    sheet_dim = calc_and_limit_sheet_dim(sheet, 0, max_row)
    row_count = sheet_dim[1]
    
    for index in range(2, row_count):
        # first lets make sure this part does not exist
        identifier = sheet.cell(row=index, column=form.order_code.data).value
        if identifier == None:
            pass
        elif find_part_in_db(ordering_code=identifier):
            flash('Duplicate identifier \"%s\" was already found in database, ignored' % identifier, "warning")
        elif already_in_list(searchfor=identifier, inlist=identifiers_to_add):
            flash('Identifier \"%s\" was found twice in import file, second instance ignored' % identifier, "warning")
        else:
            #if the part does not exist, make sure we have the package!
            package_name = sheet.cell(row=index, column=form.package.data).value
            package_id = package_in_db_get_id_or_none(name=package_name)
            if package_id == None:
                flash('no case \"%s\" found' % package_name, "error")
            else:
                flash('Part \"%s\" was succesfully added to database' % identifier, "success")    
                identifiers_to_add.append(identifier)   
                part = Part( secure_check_cell_string(sheet=sheet, row=index, column=form.part_name.data),
                             secure_check_cell_string(sheet=sheet, row=index, column=form.manufacturer.data),
                             secure_check_cell_string(sheet=sheet, row=index, column=form.order_code.data),
                             package_id,
                             secure_check_cell_string(sheet=sheet, row=index, column=form.description.data),
                            )
                obj_to_add_to_db.append(part)
    return obj_to_add_to_db

def secure_check_cell_string(sheet, column, row):
    if (column <= 0) or (row <= 0) or (column == SELECTFIELD_NONE_SELECTED):
        return "n/a"
    val = sheet.cell(row=row, column=column).value
    if val == None:
        val = "n/a"
    return val

def secure_check_cell_number(sheet, column, row):
    if (column <= 0) or (row <= 0) or (column == SELECTFIELD_NONE_SELECTED):
        return 0
    val = sheet.cell(row=row, column=column).value
    if val == None:
        val = 0
    return val


def import_packages(filepath, form):
    book = load_workbook(filename=filepath, read_only=True)
    sheet = book.active    
    obj_to_add_to_db = create_packagelist_from_package_sheet(form=form, sheet=sheet, max_row=1000)
    addToDatabase(obj_to_add_to_db)
    obj_to_add_to_db = create_alt_name_objects_from_package_sheet(form=form, sheet=sheet, max_row=1000)
    addToDatabase(obj_to_add_to_db)

def create_packagelist_from_package_sheet(form, sheet, max_row):
    obj_to_add_to_db = []
    identifiers_to_add = []
    sheet_dim = calc_and_limit_sheet_dim(sheet, 0, max_row)
    row_count = sheet_dim[1]
    
    for index in range(2, max_row):
        # first lets make sure this object does not exist in database
        identifier = sheet.cell(row=index, column=form.name.data).value
        if identifier == None:
            pass
        elif package_in_db_get_id_or_none(name=identifier) != None:
            flash('\"%s\" exists already' % identifier, "error")
        elif already_in_list(searchfor=identifier, inlist=identifiers_to_add):
            flash('Identifier \"%s\" was found twice in import file, second instance ignored' % identifier, "warning")
        else:
            flash('\"%s\" added to database' % identifier, "success")    
            identifiers_to_add.append(identifier) 
            obj = Package(  sheet.cell(row=index, column=form.name.data).value,
                            secure_check_cell_number(sheet=sheet, row=index, column=form.pin_count.data),
                            secure_check_cell_number(sheet=sheet, row=index, column=form.pitch.data),
                            secure_check_cell_number(sheet=sheet, row=index, column=form.width.data),
                            secure_check_cell_number(sheet=sheet, row=index, column=form.length.data),
                            secure_check_cell_number(sheet=sheet, row=index, column=form.height.data)
                        )
            obj_to_add_to_db.append(obj)
    return obj_to_add_to_db

def create_alt_name_objects_from_package_sheet(form, sheet, max_row):
    obj_to_add_to_db = []
    identifiers_to_add = []
    
    for index in range(2, max_row):
        identifier_str = sheet.cell(row=index, column=form.alt_package.data).value
        if identifier_str != None:
            identifier_str = identifier_str.replace(" ","")
            identifier_list = identifier_str.split(",")
            for identifier in identifier_list:
                if AltPackage.query.filter_by(name=identifier).first() != None:
                    flash('\"%s\" exists already' % identifier, "error")
                elif already_in_list(searchfor=identifier, inlist=identifiers_to_add):
                    flash('Identifier \"%s\" was found twice in import file, second instance ignored' % identifier, "warning")
                else:
                    flash('\"%s\" added to database' % identifier, "success")  
                    parent_package_name = sheet.cell(row=index, column=form.name.data).value
                    parent_id = Package.query.filter_by(name=parent_package_name).first().id
                    obj = AltPackage (
                        identifier,
                        parent_id
                    )
                    obj_to_add_to_db.append(obj)
    return obj_to_add_to_db

def calc_and_limit_sheet_dim(sheet, max_col, max_row):
    dim = sheet.calculate_dimension()
    dim = dim.split(':')
    xy = coordinate_from_string(dim[1])
    count_col = column_index_from_string(xy[0])
    count_row = xy[1]
    if max_col and (count_col > max_col):
        count_col = max_col
    if max_row and (count_row > max_row):
        count_row = max_row
    return (count_col, count_row)    

def addToDatabase(obj_to_db):
    # add all items on the list to qeue and commit
    for item in obj_to_db:
        db.session.add(item)
    db.session.commit()